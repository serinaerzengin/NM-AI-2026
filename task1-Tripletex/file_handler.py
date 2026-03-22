import base64
import csv
import io


async def process_files(files: list) -> list:
    """Process files into OpenAI-compatible content parts."""
    parts = []
    for f in files:
        filename = f.get("filename", "unknown")
        content_b64 = f.get("content_base64", "")
        mime_type = f.get("mime_type", "application/octet-stream")

        if not content_b64:
            continue

        lower = filename.lower()

        if lower.endswith(".csv"):
            try:
                raw = base64.b64decode(content_b64).decode("utf-8", errors="replace")
                reader = csv.reader(io.StringIO(raw))
                rows = list(reader)
                # Format as text table
                if rows:
                    text = f"=== CSV File: {filename} ===\n"
                    for row in rows:
                        text += " | ".join(str(c) for c in row) + "\n"
                    parts.append({"type": "input_text", "text": text})
            except Exception as e:
                parts.append({"type": "input_text", "text": f"Error reading CSV {filename}: {e}"})

        elif lower.endswith(".xlsx") or lower.endswith(".xls"):
            try:
                import openpyxl
                raw = base64.b64decode(content_b64)
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
                text = f"=== Excel File: {filename} ===\n"
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    text += f"\n--- Sheet: {sheet_name} ---\n"
                    for row in ws.iter_rows(values_only=True):
                        text += " | ".join(str(c) if c is not None else "" for c in row) + "\n"
                wb.close()
                parts.append({"type": "input_text", "text": text})
            except Exception as e:
                parts.append({"type": "input_text", "text": f"Error reading Excel {filename}: {e}"})

        elif mime_type.startswith("image/") or lower.endswith((".png", ".jpg", ".jpeg", ".gif", ".webp")):
            data_url = f"data:{mime_type};base64,{content_b64}"
            parts.append({"type": "input_image", "image_url": data_url})

        elif lower.endswith(".pdf") or mime_type == "application/pdf":
            try:
                import pdfplumber
                raw = base64.b64decode(content_b64)
                with pdfplumber.open(io.BytesIO(raw)) as pdf:
                    text = f"=== PDF File: {filename} ===\n"
                    has_text = False
                    for i, page in enumerate(pdf.pages):
                        page_text = page.extract_text() or ""
                        if page_text.strip():
                            has_text = True
                        text += f"--- Page {i+1} ---\n{page_text}\n"

                if has_text:
                    parts.append({"type": "input_text", "text": text})
                else:
                    # Scanned PDF — convert pages to images
                    import pypdfium2
                    pdf_doc = pypdfium2.PdfDocument(io.BytesIO(raw))
                    for i, page in enumerate(pdf_doc):
                        bitmap = page.render(scale=2)
                        pil_image = bitmap.to_pil()
                        img_buf = io.BytesIO()
                        pil_image.save(img_buf, format="PNG")
                        img_b64 = base64.b64encode(img_buf.getvalue()).decode()
                        parts.append({"type": "input_image", "image_url": f"data:image/png;base64,{img_b64}"})
                    pdf_doc.close()
            except Exception as e:
                parts.append({"type": "input_text", "text": f"[PDF file: {filename}, could not extract text: {e}]"})

        else:
            # Try as text
            try:
                raw = base64.b64decode(content_b64).decode("utf-8", errors="replace")
                parts.append({"type": "input_text", "text": f"=== File: {filename} ===\n{raw}"})
            except Exception:
                parts.append({"type": "input_text", "text": f"[Binary file: {filename}, cannot display]"})

    return parts
